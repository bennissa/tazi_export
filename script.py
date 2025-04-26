import openpyxl
import pandas as pd
import re
import json
from openai import OpenAI
from openpyxl.styles import NamedStyle
import os
from azure.core.credentials import AzureKeyCredential
from azure.ai.documentintelligence import DocumentIntelligenceClient
from azure.ai.documentintelligence.models import AnalyzeResult
from azure.ai.documentintelligence.models import AnalyzeDocumentRequest
import openai
from openai import OpenAI
from azure.core.credentials import AzureKeyCredential
from azure.ai.documentintelligence import DocumentIntelligenceClient
from azure.ai.documentintelligence.models import AnalyzeDocumentRequest
import openai
import json
import pandas as pd
import re
from io import StringIO
import openpyxl
import xlrd

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle

def grab_neighbour(ws, *, search_col, label, neighbour_offset=1, default=None):
    """
    Search *search_col* (1-based) for the first cell whose value contains
    *label* (case-insensitive).  Return the value in the same row,
    *search_col + neighbour_offset*.

    If the label is not found, return *default*.
    """
    label = label.lower()

    for cell in ws.iter_rows(min_col=search_col,
                             max_col=search_col,
                             values_only=False):        # get Cell objects
        cell = cell[0]
        if isinstance(cell.value, str) and label.replace(" ", "") in cell.value.lower().replace(" ", ""):
            return ws.cell(row=cell.row,
                           column=search_col + neighbour_offset).value
    return default

# Function to extract data from Excel file
# Function to extract the required data from each file
def extract_data_decharge(file_path):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    
    # Select the sheet by name
    sheet = workbook[workbook.sheetnames[0]]
    
    # Extract the required values
    # 1️⃣  Poste / bc    – stays string
    bc = str(grab_neighbour(sheet,
                            search_col=2,         # column C (1-based index)
                            label="bc :",        # or whatever header you trust
                            neighbour_offset=1    # same column, same row
                        ) or "").strip()

    # 2️⃣  Description
    description = grab_neighbour(sheet,
                                search_col=2,     # column B
                                label="description",
                                neighbour_offset=1)

    # 3️⃣  Nomenclature douanière  (digits only)
    nomenclature = re.sub(r"\D", "", str(
        grab_neighbour(sheet,
                    search_col=6,               # column D
                    label="nomenclature",
                    neighbour_offset=1) or ""
    ))

    # 4️⃣  Poids net   (formatted 0,00000 when missing)
    poids_raw = grab_neighbour(sheet,
                            search_col=6,       # column D
                            label="poids net",
                            neighbour_offset=1)
    print(poids_raw)
    poids = "{:,.5f}".format(poids_raw).replace('.', ',') if poids_raw else "0,00000"
    # Return a dictionary of extracted data with 'bc' as the key
    return {bc: {'Description': description, 'Nomenclature': nomenclature, 'Poids net': poids }}

# Function to process the uploaded file dynamically (Excel + Image)
def process_uploaded_files(image_file):
    # Extract data from Excel

    # Path to your local PDF or image file, dynamically passed
    file_path = image_file

    # Initialize Document Intelligence client for image processing
    document_intelligence_client = DocumentIntelligenceClient(endpoint=endpoint, credential=AzureKeyCredential(key))

    # Read the image file
    with open(file_path, "rb") as file:
        file_content = file.read()

        analyze_request = AnalyzeDocumentRequest(bytes_source=file_content)
        poller = document_intelligence_client.begin_analyze_document("prebuilt-layout", analyze_request)

    # Get the result from Document Intelligence
    result = poller.result(timeout=600)
    extracted_tables = []

    # Process each table from the result
    for table in result.tables:
        table_data = []
        for cell in table.cells:
            table_data.append({"row_index": cell.row_index, "column_index": cell.column_index, "content": cell.content})
        extracted_tables.append(table_data)

    return extracted_tables

# Function to clean and format the extracted data into a DataFrame
def clean_and_format_data(extracted_tables, data_combined):
    # Convert extracted tables into structured JSON for processing by OpenAI
    context = """
Vous êtes un modèle d'intelligence artificielle chargé d'extraire des informations structurées à partir de tableaux extraits de factures, packing lists ou d'autres documents nécessaires pour la déclaration en douane. Les informations extraites devront respecter un format strict, même si certaines informations sont manquantes.

Les documents peuvent être rédigés en français, anglais ou espagnol. Vous devez extraire les données de manière automatisée à partir de chaque tableau trouvé dans les documents reçus.

Votre tâche est de remplir les colonnes suivantes pour chaque produit :

- **Designation** : Description de l'article.
- **Code produit** : code numerique produit qui sont les premiers chiffres de la designation.
- **Origine** : Pays d'origine.
- **Unité** : Unité de mesure.
- **Quantité** : Quantité commandée.
- **Prix unitaire** : Prix par unité.
- **Prix total** : Prix total.
- **Poids Brut** : Poids brut.
- **Poids Net** : Poids net.
- **NC** : Nombre de carton.

En plus de ces informations par produit, vous devez également calculer et inclure les totaux suivants pour tous les produits dans le tableau :
- **Poids Brut Total** : La somme des poids bruts de tous les produits.
- **Poids Net Total** : La somme des poids nets de tous les produits.
- **Nombre de Colis Total (NC Total)** : La somme des nombres de cartons de tous les produits.

Si l'une de ces informations est manquante dans le tableau, vous devez **la laisser vide** dans le format JSON (ex : `""`).

### Important :
- Les données doivent être retournées **uniquement au format JSON**, sans texte supplémentaire ni explication.
- Pour chaque ligne du tableau, vous devez créer un objet JSON avec la structure suivante :
  - Chaque objet JSON doit contenir les mêmes clés, même si certaines valeurs sont vides.
- Vous devez extraire toutes les lignes de tous les tableaux trouvés dans le document. Si plusieurs tableaux sont présents, chaque ligne de chaque tableau doit être incluse dans le JSON.
- Ajoutez également un objet à la fin contenant les totaux pour **Poids Brut Total**, **Poids Net Total** et **Nombre de Colis Total**.

### Exemple de structure du JSON :
```json
[
  {
    "Designation": "Widget A",
    "Code produit": "8486294",
    "Origine": "Casablanca",
    "Unité": "PC",
    "Quantité": "10",
    "Prix unitaire": "500",
    "Prix total": "5000",
  },
  {
    "Poids Brut Total": "40",
    "Poids Net Total": "35",
    "NC Total": "64"
  }
]
"""

    prompt = context + json.dumps(extracted_tables)

    # Call OpenAI API to process the prompt
    completion = client.chat.completions.create(
        model="o4-mini",  # Choose a suitable GPT model
        messages=[
            {"role": "system", "content": "Tu es un assistant qui structure des données."},
            {"role": "user", "content": prompt}
        ],
    )

    # Extract and clean the JSON result
    structured_data = completion.choices[0].message.content.strip()
    cleaned_text = structured_data.replace("```json", "").replace("```", "").replace("Paix Brut", "Poids Brut").replace("Nombre unitaire", "Prix unitaire").replace("Pris", "Prix").strip()
    data = json.loads(cleaned_text)

    # Create a DataFrame from the extracted and structured data
    df = pd.json_normalize(data)
    df['Code produit'] = pd.to_numeric(df['Code produit'], errors='coerce', downcast='integer')
    df['Code produit'] = df['Code produit'].fillna(0).astype(int)

    # Map 'Code produit' to get 'Description' and 'Nomenclature' from your combined data
    df['Description'] = df['Code produit'].apply(lambda x: data_combined.get(x, {}).get('Description', ''))
    df['Nomenclature'] = df['Code produit'].apply(lambda x: data_combined.get(x, {}).get('Nomenclature', ''))
    df['Poids net'] = df['Code produit'].apply(lambda x: data_combined.get(x, {}).get('Poids net', ''))

    # Remove unnecessary columns
    df = df.drop(columns=['Poids Brut', 'Poids Net', 'NC'])
    
    # Reorganize columns
    desired_order = ['Description', 'Nomenclature', 'Designation', 'Code produit', 'Unité', 'Quantité', 'Poids net', 'Prix unitaire', 'Origine', 'Prix total', 'Poids Brut Total', 'Poids Net Total', 'NC Total']
    df = df[desired_order]

    return df

def is_only_digits(value):
    return bool(re.fullmatch(r'\d+', str(value).strip()))
def split_or_empty(text, idx, sep="/"):
    parts = str(text).split(sep, idx + 1)  # split at most idx + 1 times
    return parts[idx] if len(parts) > idx else ""

def extract_data_facture(file_path):
    ext = os.path.splitext(file_path)[-1].lower()
    products = []

    if ext == '.xls':
        # Use xlrd for .xls
        workbook = xlrd.open_workbook(file_path)
        sheet = workbook.sheet_by_index(1)  # Change to desired sheet index if needed
        row_index = 11  # Excel row 12 = index 11

        while sheet.cell_value(row_index, 2) and len(sheet.cell_value(row_index, 2).strip()) > 2:

            product = {
                "Designation": split_or_empty(sheet.cell_value(row_index, 3),0),     # Column D
                "Nomenclature": split_or_empty(sheet.cell_value(row_index, 3),1),
                "Reference": sheet.cell_value(row_index, 1),
                "Code produit": sheet.cell_value(row_index, 2),
                "Quantité": sheet.cell_value(row_index, 7),        # Column H
                "Prix total": sheet.cell_value(row_index, 9)       # Column J
            }
            products.append(product)
            row_index += 2

    elif ext == '.xlsx':
        # Use openpyxl for .xlsx
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        row_index = 12

        while sheet.cell(row=row_index, column=3).value and len(sheet.cell(row=row_index, column=3).value.strip()) > 2:
            product = {
                "Designation": split_or_empty(sheet.cell(row=row_index, column=4).value,0),
                "Nomenclature": split_or_empty(sheet.cell(row=row_index, column=4).value,1),
                "Reference": sheet.cell(row=row_index, column=2).value,
                "Code produit": sheet.cell(row=row_index, column=3).value,
                "Quantité": sheet.cell(row=row_index, column=8).value,
                "Prix total": sheet.cell(row=row_index, column=10).value
            }
            products.append(product)
            row_index += 2

    else:
        raise ValueError("Unsupported file type. Please upload .xls or .xlsx files.")
    df = pd.DataFrame(products)
    num_cols = ["Quantité", "Prix total"]

    df[num_cols] = (
        df[num_cols]
        .astype(str)                                     # ensure .str is available
        .apply(lambda col: 
            col
                .str.replace(r"\.", "", regex=True)        # remove any “.” thousands marks
                .str.replace(",", ".",  regex=False)       # turn decimal “,” → “.”
        )
        .astype(float)                                   # now safe to cast
        .fillna(0)                                       # blanks → 0
    )

    df_merged = (
        df.groupby("Reference", as_index=False)
        .agg({
            "Designation" : "first",
            "Nomenclature": "first",
            "Code produit": "first",
            "Quantité"    : "sum",
            "Prix total"  : "sum",
        })
    )

    return df_merged
def find_article(row, catalogue):
    """
    row       – a Series from df
    catalogue – your data_combined dict mapping keys → {Description, Nomenclature, Poids net, …}

    Returns (description, nomenclature, poids_net)
    """
    code = str(row.get("Code produit", "")).strip()

    # 1️⃣  exact key match
    rec = catalogue.get(code)

    # 2️⃣  fallback: find any key that contains `code`
    if rec is None and code:
        for k, v in catalogue.items():
            if code in str(k):
                rec = v
                break

    # 3️⃣  if we found something, pull the fields; otherwise empty strings
    if rec:
        return (
            rec.get("Description", ""),
            rec.get("Nomenclature", ""),
            rec.get("Poids net", "")
        )
    else:
        return "", "", ""

def process_df_facture(df,data_combined):
    df['Code produit'] = pd.to_numeric(df['Code produit'], errors='coerce', downcast='integer')
    df['Code produit'] = df['Code produit'].fillna(0).astype(int)
    # Iterate over the DataFrame and use 'Code produit' to get the 'description' and 'nomenclature' from your dictionary
    df[["Description", "Nomenclature", "Poids net"]] = (
        df.apply(find_article, axis=1, catalogue=data_combined)
        .apply(pd.Series)
    )
    return df

# Function to prepare the final DataFrame and save to Excel
def prepare_final_excel(df):
    # Read the 'ventillation_template.xlsx' Excel file
    df_template = pd.read_excel('ventillation_template.xlsx')
    print(df_template.columns)
    df_except_last = df
    # Create a DataFrame with the same number of rows as df_except_last
    df_template = pd.DataFrame(columns=df_template.columns, index=range(len(df_except_last)))


    # Fill the columns with your desired values
    df_template["Identifiant unique du fichier"] = ''
    df_template["N° ordre de l'article"] = range(1, len(df_except_last) + 1)
    df_template["Nombre Contenants"] = ''
    df_template["Type Contenant"] = "216"
    df_template["Marque (N° Envoi)"] = "TAZI EXPORT"
    df_template["Code NGP(à 10 chiffres)"] = df_except_last["Nomenclature"]
    df_template["Désignation commerciale"] = df_except_last["Description"]
    df_template["Pays d'origine"] = "MA"
    df_template["Indicateur de Paiement"] = "AP"
    df_template["Indicateur Occasion"] = "NON"
    df_template["Valeur"] = df_except_last["Prix total"]
    df_template["Devise "] = "EUR"
    df_template["Quantité Article"] = df_except_last["Quantité"]
    df_template["Unité de mesure"] = "033"
    df_template["Poids net Article"] = df_except_last["Poids net"]
    df_template["Quantité normalisée"] = df_except_last["Quantité"]
    df_template["Code Référence Accord Article"] = "UE"
    df_template["Code Référence Franchise"] = ''
    df_template["Nom et Prénom "] = ''
    df_template["CIN"] = ''

    df_template["Valeur"] = pd.to_numeric(df_template["Valeur"].replace({',': '.'}, regex=True), errors='coerce')
    df_template["Quantité Article"] = pd.to_numeric(df_template["Quantité Article"].replace({',': '.'}, regex=True), errors='coerce')
    df_template["Poids net Article"] = pd.to_numeric(df_template["Poids net Article"].replace({',': '.'}, regex=True), errors='coerce')
    df_template["Quantité normalisée"] = pd.to_numeric(df_template["Quantité normalisée"].replace({',': '.'}, regex=True), errors='coerce')
    df_template["Type Contenant"] = pd.to_numeric(df_template["Type Contenant"].replace({',': '.'}, regex=True), errors='coerce')
    df_template["Code NGP(à 10 chiffres)"] = pd.to_numeric(df_template["Code NGP(à 10 chiffres)"].replace({',': '.'}, regex=True), errors='coerce')

    # Save to Excel with openpyxl as the engine
    file_name = 'ventillation_template_populated_numeric_direct2.xlsx'

    df_grouped = df_template.groupby('Code NGP(à 10 chiffres)', as_index=False).agg(
    {
        'Identifiant unique du fichier': 'first',
        "N° ordre de l'article": 'first',
        'Nombre Contenants': 'first',
        'Type Contenant': 'first',
        'Marque (N° Envoi)': 'first',
        'Code NGP(à 10 chiffres)': 'first',  # Keep the first value for this column
        'Désignation commerciale': 'first',
        'Pays d\'origine': 'first',
        'Indicateur de Paiement': 'first',
        'Indicateur Occasion': 'first',
        'Valeur': 'sum',  # Sum 'Valeur' column
        'Devise ': 'first',
        'Quantité Article': 'sum',  # Sum 'Quantité Article' column
        'Unité de mesure': 'first',
        'Poids net Article': 'sum',  # Sum 'Poids net Article' column
        'Quantité normalisée': 'sum',  # Sum 'Quantité normalisée' column
        'Code Référence Accord Article': 'first',
        'Code Référence Franchise': 'first',
        'Nom et Prénom ': 'first',
        'CIN': 'first',
    }
    )
    print(df_grouped)   
    df_grouped.to_excel(file_name, index=False, engine='openpyxl')
    print("Excel file saved successfully!")

    return df_grouped
